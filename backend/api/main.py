import os
import tempfile
import openpyxl
from fastapi import FastAPI, UploadFile, File
from pydantic import BaseModel
from typing import Dict, Any, List
from dotenv import load_dotenv
from openai import OpenAI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse

load_dotenv()

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

interview_state: Dict[str, Any] = {}

QUESTIONS = [
    "What is the difference between a formula and a function in Excel? Give an example of each.",
    "Explain the purpose of a Pivot Table and when you would use one."
]

class ChatRequest(BaseModel):
    user_id: str = "test_user"
    message: str

def get_llm_response(messages: List[Dict[str, str]], final_prompt: bool = False, next_question: str = None):
    if final_prompt:
        system_prompt = (
            "You are a professional Excel interviewer. Review the entire conversation below. "
            "Based on the candidate's answers, provide constructive feedback and an estimated score out of 100."
        )
    else:
        system_prompt = (
            "You are a professional Excel interviewer. Acknowledge the candidate's response and ask the next question."
        )

    full_conversation = [{"role": "system", "content": system_prompt}] + messages

    try:
        completion = client.chat.completions.create(
            model="gpt-4.1-nano-2025-04-14",
            messages=full_conversation,
            max_tokens=250,
            temperature=0.6
        )
        response_content = completion.choices[0].message.content.strip()
        return response_content
    except Exception as e:
        print(f"OpenAI API error: {e}")
        return "Sorry, I'm having trouble connecting right now. Please try again later."

def create_task_excel():
    template_dir = os.path.join(os.path.dirname(__file__), 'api')
    template_path = os.path.join(template_dir, 'task_template.xlsx')

    os.makedirs(template_dir, exist_ok=True)

    if not os.path.exists(template_path):
        print(f"[INFO] Generating clean blank template at {template_path}.")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"

        # Only column headers, no data or formulas
        sheet['A1'] = "Data1"
        sheet['B1'] = "Data2"
        sheet['C1'] = "Result"

        workbook.save(template_path)
        print("[SUCCESS] Blank template created.")

    try:
        workbook = openpyxl.load_workbook(template_path)
    except Exception as e:
        print(f"[ERROR] Could not load template: {e}")
        raise e

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, "Task.xlsx")

    try:
        workbook.save(file_path)
        print(f"[SUCCESS] Task file created at: {file_path}")
    except Exception as e:
        print(f"[ERROR] Could not save task file: {e}")
        raise e

    return file_path
@app.post("/chat")
async def chat_endpoint(request_body: ChatRequest):
    user_id = request_body.user_id
    user_message = request_body.message.strip()

    # Start interview
    if user_message.lower() == "start" and user_id not in interview_state:
        interview_state[user_id] = {
            "history": [],
            "turn": 0,
            "conceptual_feedback": "",
        }
        intro_msg = "Hello! I'm your AI Excel Interviewer. I'll ask 2 questions about Excel concepts first, then provide a practical task."
        interview_state[user_id]["history"].append({"role": "assistant", "content": intro_msg})
        # Ask first question immediately
        first_question = QUESTIONS[0]
        interview_state[user_id]["history"].append({"role": "assistant", "content": first_question})
        interview_state[user_id]["turn"] = 1  # Track next turn
        return {"response": f"{intro_msg}\n\n{first_question}", "status": "ongoing"}

    state = interview_state.get(user_id)
    if not state:
        return {"response": "Please click the start button to begin the interview.", "status": "not_started"}

    # Save user answer
    state["history"].append({"role": "user", "content": user_message})

    # Ask next conceptual question
    if state["turn"] < len(QUESTIONS):
        next_question = QUESTIONS[state["turn"]]
        state["history"].append({"role": "assistant", "content": next_question})
        state["turn"] += 1
        return {"response": next_question, "status": "ongoing"}

    # If all conceptual questions answered, generate feedback and task
    if state["turn"] == len(QUESTIONS):
        # Generate LLM feedback for conceptual questions
        conceptual_feedback = get_llm_response(state["history"], final_prompt=True)
        state["conceptual_feedback"] = conceptual_feedback

        try:
            task_file_path = create_task_excel()
            print(f"Task file generated successfully: {task_file_path}")
            state["turn"] += 1  # Prevent re-generation
            return {
                "response": "Conceptual questions complete. Please download the practical task.",
                "status": "task_ready",
                "conceptual_feedback": conceptual_feedback
            }
        except Exception as e:
            print(f"Failed to create task file: {e}")
            return {"response": "Failed to generate task file. Interview cannot proceed.", "status": "error"}

    # Fallback
    return {"response": "Interview already in progress. Please download the task file.", "status": "task_ready"}

    

@app.get("/download_task")
def download_task():
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, "Task.xlsx")

    if os.path.exists(file_path):
        return FileResponse(file_path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename="Task.xlsx")
    else:
        return JSONResponse(status_code=404, content={"error": "Task file not found."})

@app.get("/check_task_file")
def check_task_file():
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, "Task.xlsx")
    exists = os.path.exists(file_path)
    return {"exists": exists}


@app.post("/upload_solution")
async def upload_solution(file: UploadFile = File(...)):
    try:
        temp_dir = tempfile.gettempdir()
        solution_path = os.path.join(temp_dir, file.filename)

        with open(solution_path, "wb") as f:
            content = await file.read()
            f.write(content)

        score, feedback = evaluate_excel_file(solution_path)

        # Add LLM-based final feedback
        user_id = "test_user"  # or track actual user
        state = interview_state.get(user_id)
        final_feedback = ""
        if state:
            # Append evaluation summary as system/user message
            state["history"].append({"role": "assistant", "content": f"Practical task scored {score}/100. Feedback: {feedback}"})
            final_feedback = get_llm_response(state["history"], final_prompt=True)
            del interview_state[user_id]

        report = f"Score: {score}/100\n\nFeedback:\n- " + "\n- ".join(feedback)
        if final_feedback:
            report += f"\n\nLLM Feedback:\n{final_feedback}"

        return {"report": report}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


def evaluate_excel_file(file_path: str):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=False)  # Important: Keep formulas
        sheet = workbook.active

        score = 100
        feedback = []

        expected_sum_formula = "=SUM(A1:A10)"
        if sheet['B2'].value and expected_sum_formula.upper() in sheet['B2'].value.upper():
            feedback.append("Correctly used SUM formula in cell B2.")
        else:
            score -= 50
            feedback.append("Missing or incorrect SUM formula in cell B2.")

        if sheet['C2'].value and ("VLOOKUP" in sheet['C2'].value.upper() or "XLOOKUP" in sheet['C2'].value.upper()):
            feedback.append("Used a lookup function in cell C2.")
        else:
            score -= 50
            feedback.append("Missing lookup function (VLOOKUP/XLOOKUP) in cell C2.")

        return max(score, 0), feedback

    except Exception as e:
        print(f"Excel evaluation error: {e}")
        return 0, ["Error evaluating Excel file."]
